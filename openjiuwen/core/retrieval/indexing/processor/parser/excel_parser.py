# coding: utf-8
# Copyright (c) Huawei Technologies Co., Ltd. 2025. All rights reserved.
"""
Excel (.xlsx), CSV, and TSV Parser

Parses Excel, CSV, and TSV files into row-wise and column-wise Documents for indexing.
Each row becomes one Document (for row retrieval); each column becomes one Document (for column retrieval).
Both are written to the same index.
"""

import asyncio
import csv
import os
from typing import TYPE_CHECKING, List, Sequence

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from openjiuwen.core.common.logging import logger
from openjiuwen.core.retrieval.common.document import Document
from openjiuwen.core.retrieval.indexing.processor.parser.auto_file_parser import register_parser
from openjiuwen.core.retrieval.indexing.processor.parser.base import Parser


def _cell_str(value) -> str:
    """Coerce cell value to string for text."""
    if value is None:
        return ""
    return str(value).strip()


def _rows_to_documents(
    rows_iter: Sequence[Sequence],
    sheet_name: str,
    base_id: str,
    sheet_index: int,
    *,
    include_header: bool = True,
) -> List[Document]:
    """Produce row docs and column docs from header row + data rows.

    If include_header is True, each resulting doc-row and doc-column is prefixed
    with its corresponding header name. If a header cell is empty, a column
    counter (e.g. Column 1, Column 2) is used instead. If include_header is False, only
    cell values are used with no header prefix.
    """
    docs: List[Document] = []
    if not rows_iter:
        return docs

    headers = [_cell_str(v) for v in rows_iter[0]]
    if not headers:
        headers = [f"Column {i+1}" for i in range(len(rows_iter[0]) or 0)]
    data_rows = rows_iter[1:]

    for row_idx, row in enumerate(data_rows, start=2):
        parts = []
        for col_idx, (h, cell_val) in enumerate(zip(headers, row)):
            if include_header and h:
                parts.append(f"{h}: {_cell_str(cell_val)}")
            elif not include_header:
                v = _cell_str(cell_val)
                if v:
                    parts.append(v)
        if not parts:
            continue
        text = ", ".join(parts)
        doc_id = f"{base_id}_s{sheet_index}_r{row_idx}"
        doc = Document(
            id_=doc_id,
            text=text,
            metadata={
                "sheet_name": sheet_name,
                "row_index": row_idx,
                "source_type": "row",
            },
        )
        docs.append(doc)

    for col_idx, col_name in enumerate(headers):
        if not col_name:
            col_name = f"Column {col_idx + 1}"
        values = []
        for row in data_rows:
            if col_idx < len(row):
                values.append(_cell_str(row[col_idx]))
        values = [v for v in values if v]
        if include_header:
            if values:
                text = f"Column name: {col_name}. Values: {', '.join(values)}"
            else:
                text = f"Column name: {col_name}. Values: (empty)"
        else:
            text = ", ".join(values) if values else ""
        doc_id = f"{base_id}_s{sheet_index}_c{col_idx}"
        doc = Document(
            id_=doc_id,
            text=text,
            metadata={
                "sheet_name": sheet_name,
                "column_name": col_name,
                "source_type": "column",
            },
        )
        docs.append(doc)

    return docs


def _sheet_to_documents(
    sheet: "Worksheet",
    sheet_name: str,
    base_id: str,
    sheet_index: int,
    *,
    include_header: bool = True,
) -> List[Document]:
    """Produce row docs and column docs for one Excel sheet."""
    rows_iter = list(sheet.iter_rows(values_only=True))
    return _rows_to_documents(
        rows_iter, sheet_name, base_id, sheet_index, include_header=include_header
    )


def _load_delimited_rows(path: str, delimiter: str = ",") -> List[List[str]]:
    """Read delimited text file (CSV, TSV, etc.) into list of rows (first row = headers)."""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f, delimiter=delimiter))


@register_parser([".xlsx", ".XLSX", ".csv", ".CSV", ".tsv", ".TSV"])
class ExcelParser(Parser):
    """
    Parser for Excel (.xlsx), CSV, and TSV files.
    Emits both row-wise and column-wise Documents for the same index.
    """

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

    async def parse(self, doc: str, doc_id: str = "", **kwargs) -> List[Document]:
        """
        Parse an Excel, CSV, or TSV file into row and column Documents.

        Args:
            doc: Path to .xlsx, .csv, or .tsv file.
            doc_id: Optional base document ID (used as prefix for each doc).
            include_header: If True (default), prepend column name to each cell in row
                docs and to column doc text; if False, use only cell values.

        Returns:
            List of Documents (row docs + column docs for each sheet / file).
        """
        from openjiuwen.core.common.exception.codes import StatusCode
        from openjiuwen.core.common.exception.errors import build_error

        if not os.path.exists(doc):
            raise build_error(
                StatusCode.RETRIEVAL_INDEXING_FILE_NOT_FOUND,
                error_msg=f"File {doc} does not exist",
            )

        ext = os.path.splitext(doc)[-1].lower()
        base_id = doc_id or doc or "tabular"

        if ext in (".csv", ".tsv"):
            delimiter = "\t" if ext == ".tsv" else ","
            try:
                rows = await asyncio.to_thread(_load_delimited_rows, doc, delimiter)
            except Exception as e:
                logger.exception("Failed to parse %s %s: %s", ext.upper().lstrip("."), doc, e)
                raise build_error(
                    StatusCode.RETRIEVAL_INDEXING_FORMAT_NOT_SUPPORT,
                    error_msg=f"Parse failed for {doc}: {e}",
                    cause=e,
                ) from e
            sheet_name = os.path.basename(doc) or "default"
            include_header = kwargs.get("include_header", True)
            documents = _rows_to_documents(
                rows, sheet_name, base_id, 0, include_header=include_header
            )
        else:
            include_header = kwargs.get("include_header", True)

            def _load_and_parse():
                from openpyxl import load_workbook

                wb = load_workbook(doc, read_only=True, data_only=True)
                try:
                    out = []
                    for si, sheet in enumerate(wb.worksheets):
                        for d in _sheet_to_documents(
                            sheet, sheet.title, base_id, si, include_header=include_header
                        ):
                            out.append(d)
                    return out
                finally:
                    wb.close()

            try:
                documents = await asyncio.to_thread(_load_and_parse)
            except Exception as e:
                logger.exception("Failed to parse Excel %s: %s", doc, e)
                raise build_error(
                    StatusCode.RETRIEVAL_INDEXING_FORMAT_NOT_SUPPORT,
                    error_msg=f"Excel parse failed for {doc}: {e}",
                    cause=e,
                ) from e

        logger.info("Parsed %s: %d documents (rows + columns)", doc, len(documents))
        return documents

    def supports(self, doc: str) -> bool:
        """True for .xlsx, .csv, and .tsv file paths."""
        if not doc:
            return False
        ext = os.path.splitext(doc)[-1].lower()
        return ext in (".xlsx", ".csv", ".tsv")
