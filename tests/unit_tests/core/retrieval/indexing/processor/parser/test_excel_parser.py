# Copyright (c) Huawei Technologies Co., Ltd. 2025-2026. All rights reserved.
"""
Excel parser test cases
"""

import os
import tempfile

import pytest

from openjiuwen.core.retrieval.common.document import Document
from openjiuwen.core.retrieval.indexing.processor.parser.excel_parser import (
    ExcelParser,
    _cell_str,
    _rows_to_documents,
)


def _make_sample_xlsx(path: str) -> None:
    """Create a small .xlsx with one sheet: headers Name, Dept, Sales; two data rows."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Dept", "Sales"])
    ws.append(["Alice", "Sales", 100])
    ws.append(["Bob", "Tech", 200])
    wb.save(path)
    wb.close()


def _make_multi_sheet_xlsx(path: str) -> None:
    """Create .xlsx with two sheets."""
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "People"
    ws1.append(["Name", "Age"])
    ws1.append(["Alice", 30])

    ws2 = wb.create_sheet("Products")
    ws2.append(["Item", "Price"])
    ws2.append(["Pen", 5])
    ws2.append(["Book", 20])
    wb.save(path)
    wb.close()


def _make_header_only_xlsx(path: str) -> None:
    """Create .xlsx with only a header row (no data)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    ws.append(["Col1", "Col2", "Col3"])
    wb.save(path)
    wb.close()


def _make_empty_header_xlsx(path: str) -> None:
    """Create .xlsx with empty/None header cells."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([None, "Score", ""])
    ws.append(["Alice", 90, "pass"])
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# _cell_str helper
# ---------------------------------------------------------------------------
class TestCellStr:
    """Tests for _cell_str utility"""

    @staticmethod
    def test_none():
        assert _cell_str(None) == ""

    @staticmethod
    def test_string():
        assert _cell_str("  hello  ") == "hello"

    @staticmethod
    def test_int():
        assert _cell_str(100) == "100"

    @staticmethod
    def test_float():
        assert _cell_str(3.14) == "3.14"

    @staticmethod
    def test_empty_string():
        assert _cell_str("") == ""


# ---------------------------------------------------------------------------
# _rows_to_documents
# ---------------------------------------------------------------------------
class TestRowsToDocuments:
    """Tests for _rows_to_documents internal function"""

    @staticmethod
    def test_empty_input():
        assert _rows_to_documents([], "Sheet1", "base", 0) == []

    @staticmethod
    def test_header_only_no_data():
        """Header-only table: should produce column docs but no row docs."""
        rows = [["A", "B", "C"]]
        docs = _rows_to_documents(rows, "Sheet1", "base", 0)
        row_docs = [d for d in docs if d.metadata["source_type"] == "row"]
        col_docs = [d for d in docs if d.metadata["source_type"] == "column"]
        assert len(row_docs) == 0
        assert len(col_docs) == 3

    @staticmethod
    def test_row_text_format():
        """Row text should be 'Header: Value' pairs joined by comma."""
        rows = [["Name", "Age"], ["Alice", "30"]]
        docs = _rows_to_documents(rows, "S", "b", 0)
        row_doc = next(d for d in docs if d.metadata["source_type"] == "row")
        assert row_doc.text == "Name: Alice, Age: 30"

    @staticmethod
    def test_column_text_format():
        """Column text should list all values."""
        rows = [["City"], ["Beijing"], ["Shanghai"]]
        docs = _rows_to_documents(rows, "S", "b", 0)
        col_doc = next(d for d in docs if d.metadata["source_type"] == "column")
        assert "Column name: City" in col_doc.text
        assert "Beijing" in col_doc.text
        assert "Shanghai" in col_doc.text

    @staticmethod
    def test_doc_ids():
        """Doc IDs should follow base_s{sheet}_r{row} / base_s{sheet}_c{col} pattern."""
        rows = [["X"], ["v1"]]
        docs = _rows_to_documents(rows, "S", "mybase", 2)
        ids = {d.id_ for d in docs}
        assert "mybase_s2_r2" in ids
        assert "mybase_s2_c0" in ids

    @staticmethod
    def test_include_header_false():
        """With include_header=False, row/column text has no header prefix."""
        rows = [["Name", "Age"], ["Alice", "30"]]
        docs = _rows_to_documents(rows, "S", "b", 0, include_header=False)
        row_doc = next(d for d in docs if d.metadata["source_type"] == "row")
        assert row_doc.text == "Alice, 30"
        col_doc = next(d for d in docs if d.metadata["source_type"] == "column")
        assert "Column name:" not in col_doc.text
        assert "Alice" in col_doc.text or "30" in col_doc.text


# ---------------------------------------------------------------------------
# ExcelParser.supports
# ---------------------------------------------------------------------------
class TestExcelParserSupports:
    """ExcelParser.supports() tests"""

    @staticmethod
    def test_supports_xlsx():
        parser = ExcelParser()
        assert parser.supports("/some/file.xlsx") is True
        assert parser.supports("/some/file.XLSX") is True

    @staticmethod
    def test_supports_csv():
        parser = ExcelParser()
        assert parser.supports("/some/file.csv") is True
        assert parser.supports("/some/file.CSV") is True

    @staticmethod
    def test_supports_tsv():
        parser = ExcelParser()
        assert parser.supports("/some/file.tsv") is True
        assert parser.supports("/some/file.TSV") is True

    @staticmethod
    def test_rejects_other_extensions():
        parser = ExcelParser()
        assert parser.supports("/some/file.pdf") is False
        assert parser.supports("/some/file.xls") is False
        assert parser.supports("/some/file.txt") is False

    @staticmethod
    def test_rejects_empty():
        parser = ExcelParser()
        assert parser.supports("") is False
        assert parser.supports(None) is False


# ---------------------------------------------------------------------------
# ExcelParser.parse — xlsx
# ---------------------------------------------------------------------------
class TestExcelParserXlsx:
    """ExcelParser.parse() tests for .xlsx files"""

    @pytest.mark.asyncio
    async def test_parse_row_and_column_docs(self):
        """Standard xlsx: 2 row docs + 3 column docs"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _make_sample_xlsx(path)
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="excel1")

            row_docs = [d for d in docs if d.metadata.get("source_type") == "row"]
            col_docs = [d for d in docs if d.metadata.get("source_type") == "column"]
            assert len(row_docs) == 2
            assert len(col_docs) == 3

            row_docs.sort(key=lambda d: d.metadata["row_index"])
            assert row_docs[0].metadata["sheet_name"] == "Sheet1"
            assert row_docs[0].metadata["row_index"] == 2
            assert "Alice" in row_docs[0].text and "Sales" in row_docs[0].text
            assert row_docs[1].metadata["row_index"] == 3
            assert "Bob" in row_docs[1].text and "Tech" in row_docs[1].text

            col_names = {d.metadata["column_name"] for d in col_docs}
            assert col_names == {"Name", "Dept", "Sales"}
            name_col = next(d for d in col_docs if d.metadata["column_name"] == "Name")
            assert "Alice" in name_col.text and "Bob" in name_col.text
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_doc_id_prefix(self):
        """doc_id is used as ID prefix"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _make_sample_xlsx(path)
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="my_id")
            assert all(d.id_.startswith("my_id") for d in docs)
            row_one = next(
                d for d in docs
                if d.metadata.get("source_type") == "row" and d.metadata.get("row_index") == 2
            )
            assert row_one.id_ == "my_id_s0_r2"
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_multi_sheet(self):
        """Multi-sheet xlsx: docs from all sheets"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _make_multi_sheet_xlsx(path)
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="ms")

            sheets = {d.metadata["sheet_name"] for d in docs}
            assert "People" in sheets
            assert "Products" in sheets

            people_rows = [
                d for d in docs
                if d.metadata["sheet_name"] == "People" and d.metadata["source_type"] == "row"
            ]
            product_rows = [
                d for d in docs
                if d.metadata["sheet_name"] == "Products" and d.metadata["source_type"] == "row"
            ]
            assert len(people_rows) == 1
            assert len(product_rows) == 2
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_header_only(self):
        """Header-only xlsx: column docs but no row docs"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _make_header_only_xlsx(path)
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="hdr")

            row_docs = [d for d in docs if d.metadata["source_type"] == "row"]
            col_docs = [d for d in docs if d.metadata["source_type"] == "column"]
            assert len(row_docs) == 0
            assert len(col_docs) == 3
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_empty_header_fallback(self):
        """Empty header cells should fallback to 'Column N' naming"""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            _make_empty_header_xlsx(path)
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="eh")
            col_docs = [d for d in docs if d.metadata["source_type"] == "column"]
            col_names = {d.metadata["column_name"] for d in col_docs}
            assert "Score" in col_names
            # Empty headers should get fallback names
            fallback_names = col_names - {"Score"}
            assert len(fallback_names) >= 1
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_missing_file_raises(self):
        from openjiuwen.core.common.exception.errors import BaseError

        parser = ExcelParser()
        with pytest.raises(BaseError, match="does not exist"):
            await parser.parse("/nonexistent/file.xlsx", doc_id="x")


# ---------------------------------------------------------------------------
# ExcelParser.parse — csv
# ---------------------------------------------------------------------------
class TestExcelParserCsv:
    """ExcelParser.parse() tests for .csv files"""

    @pytest.mark.asyncio
    async def test_parse_csv_row_and_column_docs(self):
        """CSV: same row/column doc shape as xlsx"""
        with tempfile.NamedTemporaryFile(
            suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8"
        ) as f:
            path = f.name
            f.write("Name,Dept,Sales\nAlice,Sales,100\nBob,Tech,200\n")
        try:
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="csv1")
            row_docs = [d for d in docs if d.metadata.get("source_type") == "row"]
            col_docs = [d for d in docs if d.metadata.get("source_type") == "column"]
            assert len(row_docs) == 2
            assert len(col_docs) == 3
            assert row_docs[0].metadata["sheet_name"] == os.path.basename(path)
            assert "Alice" in row_docs[0].text
            col_names = {d.metadata["column_name"] for d in col_docs}
            assert col_names == {"Name", "Dept", "Sales"}
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_csv_with_empty_cells(self):
        """CSV with empty cells should not crash"""
        with tempfile.NamedTemporaryFile(
            suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8"
        ) as f:
            path = f.name
            f.write("A,B\n1,\n,3\n")
        try:
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="csv_empty")
            row_docs = [d for d in docs if d.metadata["source_type"] == "row"]
            assert len(row_docs) == 2
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_csv_header_only(self):
        """CSV with only headers: column docs but no row docs"""
        with tempfile.NamedTemporaryFile(
            suffix=".csv", delete=False, mode="w", newline="", encoding="utf-8"
        ) as f:
            path = f.name
            f.write("X,Y,Z\n")
        try:
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="csv_hdr")
            row_docs = [d for d in docs if d.metadata["source_type"] == "row"]
            col_docs = [d for d in docs if d.metadata["source_type"] == "column"]
            assert len(row_docs) == 0
            assert len(col_docs) == 3
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_csv_missing_file_raises(self):
        from openjiuwen.core.common.exception.errors import BaseError

        parser = ExcelParser()
        with pytest.raises(BaseError, match="does not exist"):
            await parser.parse("/nonexistent/file.csv", doc_id="x")


# ---------------------------------------------------------------------------
# ExcelParser.parse — tsv
# ---------------------------------------------------------------------------
class TestExcelParserTsv:
    """ExcelParser.parse() tests for .tsv files"""

    @pytest.mark.asyncio
    async def test_parse_tsv_row_and_column_docs(self):
        """TSV: same row/column doc shape as CSV (tab delimiter)."""
        with tempfile.NamedTemporaryFile(
            suffix=".tsv", delete=False, mode="w", newline="", encoding="utf-8"
        ) as f:
            path = f.name
            f.write("Name\tDept\tSales\nAlice\tSales\t100\nBob\tTech\t200\n")
        try:
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="tsv1")
            row_docs = [d for d in docs if d.metadata.get("source_type") == "row"]
            col_docs = [d for d in docs if d.metadata.get("source_type") == "column"]
            assert len(row_docs) == 2
            assert len(col_docs) == 3
            assert "Alice" in row_docs[0].text
            col_names = {d.metadata["column_name"] for d in col_docs}
            assert col_names == {"Name", "Dept", "Sales"}
        finally:
            os.unlink(path)

    @pytest.mark.asyncio
    async def test_parse_tsv_include_header_false(self):
        """TSV with include_header=False: no header prefix in text."""
        with tempfile.NamedTemporaryFile(
            suffix=".tsv", delete=False, mode="w", newline="", encoding="utf-8"
        ) as f:
            path = f.name
            f.write("A\tB\n1\t2\n")
        try:
            parser = ExcelParser()
            docs = await parser.parse(path, doc_id="tsv2", include_header=False)
            row_docs = [d for d in docs if d.metadata.get("source_type") == "row"]
            col_docs = [d for d in docs if d.metadata.get("source_type") == "column"]
            assert len(row_docs) == 1
            assert row_docs[0].text == "1, 2"
            assert len(col_docs) == 2
            # Column docs without header: just values
            col_texts = [d.text for d in col_docs]
            assert "1" in col_texts[0] and "2" in col_texts[1]
        finally:
            os.unlink(path)
